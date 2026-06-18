#!/usr/bin/env python3
"""
generate_report.py — Aggregate Maven Surefire XML reports into an Excel workbook.

Usage:
    python generate_report.py --surefire target/surefire-reports --out backend-test-report.xlsx
"""
import argparse
import datetime
import glob
import os
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl not installed — run: pip install openpyxl")
    sys.exit(1)

# ── Colours ──────────────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
BLUE   = PatternFill("solid", fgColor="BDD7EE")
GREY   = PatternFill("solid", fgColor="D9D9D9")
ORANGE = PatternFill("solid", fgColor="F4B942")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1B5E20")

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

SECURITY_TESTS = {
    "[V-01]", "[V-02]", "[V-03]", "[V-04]", "[V-05]", "[V-06]",
    "[V-07]", "[V-08]", "[V-09]", "[V-10]", "[V-11]", "[V-12]",
    "[V-13]", "[V-14]", "[V-15]", "[V-16]", "[V-17]", "[V-18]",
    "[SECURITY]",
}


def is_security(name: str) -> bool:
    return any(tag in name for tag in SECURITY_TESTS)


def _border(cell, fill=None):
    cell.border = THIN
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill


def _header_row(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[1].height = 24


def _autowidth(ws, max_w=70):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, max_w)


# ── Parser ────────────────────────────────────────────────────────────────────

def parse_surefire(directory: str):
    tests = []
    for xml_path in sorted(glob.glob(os.path.join(directory, "TEST-*.xml"))):
        tree = ET.parse(xml_path)
        root = tree.getroot()
        suite = root.get("name", os.path.basename(xml_path))
        # Shorten class name for display
        short_suite = suite.split(".")[-1]

        for tc in root.iter("testcase"):
            name = tc.get("name", "unknown")
            classname = tc.get("classname", suite)
            duration_s = float(tc.get("time", 0) or 0)
            failure = tc.find("failure")
            error_el = tc.find("error")
            skip_el = tc.find("skipped")

            if failure is not None:
                status = "FAILED"
                err_msg = (failure.get("message") or "") + "\n" + (failure.text or "")
            elif error_el is not None:
                status = "ERROR"
                err_msg = (error_el.get("message") or "") + "\n" + (error_el.text or "")
            elif skip_el is not None:
                status = "SKIPPED"
                err_msg = skip_el.get("message", "")
            else:
                status = "PASSED"
                err_msg = ""

            tests.append({
                "suite": short_suite,
                "full_class": classname,
                "name": name,
                "status": status,
                "duration_ms": round(duration_s * 1000),
                "error": err_msg.strip()[:600],
                "security": is_security(name),
                "category": _category(short_suite),
            })
    return tests


def _category(suite_name: str) -> str:
    if "Security" in suite_name or "RateLimit" in suite_name:
        return "Security & Vulnerability"
    if "Jwt" in suite_name:
        return "JWT / Auth"
    if "Disease" in suite_name or "Farming" in suite_name or "Chat" in suite_name or "Grok" in suite_name:
        return "AI / Farming"
    if "Exception" in suite_name:
        return "Exception Handling"
    return "General"


def _status_fill(status: str) -> PatternFill:
    if status == "PASSED": return GREEN
    if status in ("FAILED", "ERROR"): return RED
    if status == "SKIPPED": return YELLOW
    return GREY


# ── Sheets ────────────────────────────────────────────────────────────────────

def build_summary(wb, tests, timestamp):
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    total   = len(tests)
    passed  = sum(1 for t in tests if t["status"] == "PASSED")
    failed  = sum(1 for t in tests if t["status"] in ("FAILED", "ERROR"))
    skipped = sum(1 for t in tests if t["status"] == "SKIPPED")
    rate    = f"{(passed / total * 100):.1f}%" if total else "N/A"
    security_total  = sum(1 for t in tests if t["security"])
    security_passed = sum(1 for t in tests if t["security"] and t["status"] == "PASSED")

    def row(label, value, fill=None):
        ws.append([label, value])
        r = ws.max_row
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 1).border = THIN
        ws.cell(r, 1).alignment = Alignment(vertical="center")
        ws.cell(r, 2).border = THIN
        ws.cell(r, 2).alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            ws.cell(r, 2).fill = fill

    # Title
    ws.append(["AgriDirect Backend — Test Report"])
    ws["A1"].font = Font(bold=True, size=16, color="1B5E20")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.append(["Generated", timestamp])
    ws.append([])

    row("Total Tests", total)
    row("Passed", passed, GREEN)
    row("Failed / Error", failed, RED if failed else GREEN)
    row("Skipped", skipped, YELLOW if skipped else GREY)
    row("Pass Rate", rate, GREEN if failed == 0 else RED)
    ws.append([])
    row("Security Tests", security_total)
    row("Security Passed", security_passed, GREEN if security_passed == security_total else RED)
    ws.append([])

    # By category
    ws.append(["Category Breakdown", ""])
    ws.cell(ws.max_row, 1).font = HEADER_FONT
    ws.cell(ws.max_row, 1).fill = HEADER_FILL
    ws.cell(ws.max_row, 2).fill = HEADER_FILL

    categories = {}
    for t in tests:
        cat = t["category"]
        categories.setdefault(cat, {"passed": 0, "failed": 0, "total": 0})
        categories[cat]["total"] += 1
        if t["status"] == "PASSED":
            categories[cat]["passed"] += 1
        else:
            categories[cat]["failed"] += 1

    for cat, counts in sorted(categories.items()):
        fill = GREEN if counts["failed"] == 0 else RED
        row(cat, f"{counts['passed']}/{counts['total']}", fill)


def build_details(wb, tests):
    ws = wb.create_sheet("All Tests")
    _header_row(ws, ["#", "Category", "Suite", "Test Name", "Status", "Duration (ms)", "Security?"])
    for i, t in enumerate(tests, 1):
        ws.append([
            i, t["category"], t["suite"], t["name"], t["status"],
            t["duration_ms"], "Yes" if t["security"] else "",
        ])
        fill = _status_fill(t["status"])
        sec_fill = BLUE if t["security"] else fill
        for col in range(1, 7):
            _border(ws.cell(i + 1, col), fill)
        _border(ws.cell(i + 1, 7), sec_fill)
        if t["security"]:
            ws.cell(i + 1, 7).value = "Yes"
    _autowidth(ws)


def build_security(wb, tests):
    ws = wb.create_sheet("Security Tests")
    sec_tests = [t for t in tests if t["security"]]
    if not sec_tests:
        ws.append(["No security-tagged tests found."])
        return

    _header_row(ws, ["#", "Suite", "Test Name", "Status", "Duration (ms)", "Error"])
    for i, t in enumerate(sec_tests, 1):
        ws.append([i, t["suite"], t["name"], t["status"], t["duration_ms"], t["error"]])
        fill = _status_fill(t["status"])
        for col in range(1, 7):
            _border(ws.cell(i + 1, col), fill)
    _autowidth(ws)

    total  = len(sec_tests)
    passed = sum(1 for t in sec_tests if t["status"] == "PASSED")
    ws.append([])
    ws.append(["", "", f"Security pass rate: {passed}/{total} ({passed/total*100:.0f}%)" if total else "N/A"])
    ws.cell(ws.max_row, 3).font = Font(bold=True, size=13)
    ws.cell(ws.max_row, 3).fill = GREEN if passed == total else RED


def build_failures(wb, tests):
    ws = wb.create_sheet("Failures")
    failed = [t for t in tests if t["status"] in ("FAILED", "ERROR")]

    if not failed:
        ws["A1"].value = "All 102 tests passed!"
        ws["A1"].font = Font(bold=True, size=16, color="1B5E20")
        return

    _header_row(ws, ["#", "Suite", "Test Name", "Status", "Error"])
    for i, t in enumerate(failed, 1):
        ws.append([i, t["suite"], t["name"], t["status"], t["error"]])
        for col in range(1, 6):
            _border(ws.cell(i + 1, col), RED)
    _autowidth(ws)


def build_by_suite(wb, tests):
    ws = wb.create_sheet("By Suite")
    _header_row(ws, ["Suite", "Category", "Total", "Passed", "Failed", "Pass Rate"])

    suites = {}
    for t in tests:
        k = (t["suite"], t["category"])
        suites.setdefault(k, {"passed": 0, "failed": 0, "total": 0})
        suites[k]["total"] += 1
        if t["status"] == "PASSED":
            suites[k]["passed"] += 1
        else:
            suites[k]["failed"] += 1

    for (suite, cat), c in sorted(suites.items()):
        rate = f"{c['passed']/c['total']*100:.0f}%"
        fill = GREEN if c["failed"] == 0 else RED
        ws.append([suite, cat, c["total"], c["passed"], c["failed"], rate])
        for col in range(1, 7):
            _border(ws.cell(ws.max_row, col), fill)
    _autowidth(ws)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--surefire", default="target/surefire-reports",
                   help="Directory containing TEST-*.xml files")
    p.add_argument("--out", default=None, help="Output .xlsx path")
    args = p.parse_args()

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ts_file   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = args.out or f"backend-test-report-{ts_file}.xlsx"

    print(f"[report] Reading surefire reports from: {args.surefire}")
    tests = parse_surefire(args.surefire)

    if not tests:
        print("[report] No test results found. Run 'mvn test' first.")
        sys.exit(1)

    total  = len(tests)
    passed = sum(1 for t in tests if t["status"] == "PASSED")
    failed = sum(1 for t in tests if t["status"] in ("FAILED", "ERROR"))

    print(f"[report] {total} tests: {passed} passed, {failed} failed")

    wb = openpyxl.Workbook()
    build_summary(wb, tests, timestamp)
    build_details(wb, tests)
    build_security(wb, tests)
    build_failures(wb, tests)
    build_by_suite(wb, tests)

    wb.save(out_path)
    print(f"[report] Saved: {out_path}")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
